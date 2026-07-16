import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
import sys

# ============================================================
# 解除 Pillow 图片大小限制
# ============================================================
from PIL import Image
Image.MAX_IMAGE_PIXELS = None

# 后续代码...

# ============================================================
# 配置参数
# ============================================================

def parse_folder_name(folder_name):
    """
    解析文件夹名称，提取：专案、阶段、Config、机台号、M/PBO
    格式：专案 阶段 Config 机台号 M/PBO
    例如：ATW-E C3.0 ConfigA 123 MBO
    """
    parts = folder_name.strip().split()
    if len(parts) < 5:
        print(f"⚠️ 警告：文件夹名称格式不正确，需要至少5个部分")
        print(f"   当前：{folder_name}")
        return None
    
    return {
        "专案": parts[0],
        "阶段": parts[1],
        "Config": parts[2],
        "机台号": parts[3],
        "报告类型": parts[4]  # MBO 或 PBO
    }


def extract_revision_from_file(file_path, search_pattern):
    """
    从文件中搜索并提取版本号
    """
    if not file_path or not os.path.exists(file_path):
        return None
    
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            
        match = re.search(search_pattern, content)
        if match:
            return match.group(1)
        return None
    except Exception as e:
        print(f"   ⚠️ 读取文件失败：{e}")
        return None


def get_first_file_in_folder(folder_path):
    """
    获取文件夹中的第一个有效文件（排除临时文件）
    """
    if not os.path.exists(folder_path):
        return None
    
    files = [
        f for f in os.listdir(folder_path) 
        if os.path.isfile(os.path.join(folder_path, f))
        and not f.startswith('~$')  # 排除 Excel 临时文件
        and not f.startswith('.')    # 排除隐藏文件
    ]
    
    if files:
        return os.path.join(folder_path, files[0])
    return None


def extract_cx_x_from_string(text):
    """
    从字符串中提取并转换 Cxxxx 格式为 Cx.x
    例如: "C4160 PBO" → "C4.1"
          "C4160" → "C4.1"
          "C1234" → "C1.2"
    """
    if not text:
        return ""
    
    text_str = str(text)
    
    # 匹配 C 开头后跟数字的模式
    match = re.search(r'(C)(\d+)', text_str)
    if match:
        c_prefix = match.group(1)  # "C"
        digits = match.group(2)    # "4160"
        if len(digits) >= 2:
            return f"{c_prefix}{digits[0]}.{digits[1]}"
        else:
            return f"{c_prefix}{digits}"
    
    # 如果没有匹配到，返回原始字符串（或空）
    return text_str.split()[0] if text_str else ""

def update_x_section_report(wb, data_folder_path, selected_fixtures):
    """
    更新 X-Section Sheet
    1. 清空未选中 Fixture 的 E-Q 列（文本和图片）
    2. 对选中的 Fixture：清空 H-I 图片和 J-M 数值
    3. 从 OCR CSV 按顺序读取数据，每12行对应一个 Fixture
    4. 插入图片（子文件夹 1 → H 列，子文件夹 2 → I 列）
    """
    print("\n" + "=" * 60)
    print("📝 更新 X-Section Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过更新")
        return

    print(f"   📌 用户选择的 Fixture: {selected_fixtures}")
    print(f"   📌 每个 Fixture 固定 12 行（3 Samples × 4 Positions）")
    print(f"   📌 OCR 数据按顺序读取：前12行→第一个Fixture，接下来12行→第二个Fixture...")

    # ---- 目标行范围（每个Fixture 12行） ----
    target_ranges = {
        1: (10, 21), 2: (22, 33), 3: (34, 45), 4: (46, 57),
        5: (58, 69), 6: (70, 81), 7: (82, 93), 8: (94, 105),
    }

    # ---- 获取目标 Sheet ----
    try:
        ws_target = wb["X-Section"]
    except KeyError:
        print("   ⚠️ 未找到 'X-Section' Sheet，跳过更新")
        return

    # ---- 合并单元格安全赋值 ----
    def get_merged_cell_top_left(ws, row, col):
        for merged_range in ws.merged_cells.ranges:
            if row in range(merged_range.min_row, merged_range.max_row + 1) and col in range(merged_range.min_col, merged_range.max_col + 1):
                return merged_range.min_row, merged_range.min_col
        return None, None

    def safe_set_cell_value(ws, row, col, value=None):
        top_row, top_col = get_merged_cell_top_left(ws, row, col)
        if top_row is not None:
            ws.cell(row=top_row, column=top_col).value = value
        else:
            ws.cell(row=row, column=col).value = value

    # ============================================================
    # 1. 清空旧图片
    # ============================================================
    if ws_target._images:
        ws_target._images.clear()
        print("   ✅ 已移除所有旧图片")

    # ============================================================
    # 2. 清空未选中的 Fixture 的 E-Q 列
    # ============================================================
    print("📝 清空未选中 Fixture 的 E-Q 列（文本）...")
    for fixture_num in range(1, 9):
        if fixture_num in selected_fixtures:
            continue
        start_row, end_row = target_ranges[fixture_num]
        for row in range(start_row, end_row + 1):
            for col in range(5, 18):
                safe_set_cell_value(ws_target, row, col, None)
    print("   ✅ 清空完成")

    # ============================================================
    # 3. 处理选中的 Fixture：清空 H-I 和 J-M 列
    # ============================================================
    print("📝 清空选中 Fixture 的 H-I 列和 J-M 列...")
    for fixture_num in selected_fixtures:
        start_row, end_row = target_ranges[fixture_num]
        for row in range(start_row, end_row + 1):
            safe_set_cell_value(ws_target, row, 8, None)   # H
            safe_set_cell_value(ws_target, row, 9, None)   # I
            for col in range(10, 15):
                safe_set_cell_value(ws_target, row, col, None)
    print("   ✅ 清空完成")

    # ============================================================
    # 4. 读取 OCR CSV 数据（按顺序存储）
    # ============================================================
    xsection_folder = os.path.join(data_folder_path, "X-section")
    if not os.path.exists(xsection_folder):
        print(f"   ⚠️ X-section 文件夹不存在：{xsection_folder}")
        return

    csv_files = [f for f in os.listdir(xsection_folder) if f.lower().endswith('.csv')]
    if not csv_files:
        print(f"   ⚠️ X-section 文件夹中没有 CSV 文件")
        return
    ocr_csv_path = os.path.join(xsection_folder, csv_files[0])
    print(f"   📄 读取 OCR 文件: {csv_files[0]}")

    import csv
    all_rows = []  # 存储所有行数据（每个元素是一个dict，包含映射后的列值）
    with open(ocr_csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        print(f"   📋 CSV 列名: {reader.fieldnames}")
        for row in reader:
            # 提取四个数值（列名可能是 No.2, No.4, No.5, No.3）
            values = {}
            for key in ['No.2', 'No.4', 'No.5', 'No.3']:
                if key in row:
                    val_str = row.get(key, '').strip()
                    if val_str and val_str != 'OCR失败':
                        try:
                            values[key] = float(val_str)
                        except:
                            pass
            # 根据数值范围映射到 J, K, L, M
            mapped = {}
            for key, val in values.items():
                if 91 <= val <= 139:
                    mapped['J'] = val
                elif 150 <= val <= 170:
                    mapped['K'] = val
                elif 20 <= val <= 90:
                    mapped['L'] = val
                elif val < 15:
                    mapped['M'] = val
            all_rows.append(mapped)
    
    total_ocr_rows = len(all_rows)
    print(f"   ✅ 读取到 {total_ocr_rows} 行 OCR 数据")
    if total_ocr_rows == 0:
        print("   ⚠️ 没有有效的 OCR 数据")
        return
    
    # 按 Fixture 分组：每12行一组
    rows_per_fixture = 12
    fixture_data_groups = []
    for i in range(0, total_ocr_rows, rows_per_fixture):
        group = all_rows[i:i+rows_per_fixture]
        if len(group) == rows_per_fixture:
            fixture_data_groups.append(group)
        else:
            print(f"   ⚠️ 最后一组数据不足 {rows_per_fixture} 行（{len(group)}行），已跳过")
    
    print(f"   ✅ 数据分组为 {len(fixture_data_groups)} 个 Fixture（每组 {rows_per_fixture} 行）")
    if len(fixture_data_groups) < len(selected_fixtures):
        print(f"   ⚠️ 数据组数量 ({len(fixture_data_groups)}) 少于选中的 Fixture 数量 ({len(selected_fixtures)})")

    # ============================================================
    # 5. 插入图片和填充数据（按顺序）
    # ============================================================
    folder1 = os.path.join(xsection_folder, "1")
    folder2 = os.path.join(xsection_folder, "2")
    if not os.path.exists(folder1) or not os.path.exists(folder2):
        print(f"   ⚠️ 子文件夹不存在")
        return

    image_extensions = ['.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff', '.tif']
    def get_image_files(folder_path):
        files = []
        for f in os.listdir(folder_path):
            if os.path.isfile(os.path.join(folder_path, f)):
                ext = os.path.splitext(f)[1].lower()
                if ext in image_extensions:
                    files.append(f)
        files.sort()
        return files

    images1 = get_image_files(folder1)
    images2 = get_image_files(folder2)
    print(f"   📷 子文件夹 '1' 中找到 {len(images1)} 张图片")
    print(f"   📷 子文件夹 '2' 中找到 {len(images2)} 张图片")

    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from io import BytesIO

    def insert_image_to_cell(ws, row, col, image_path):
        try:
            with open(image_path, 'rb') as f:
                img_data = f.read()
            img_stream = BytesIO(img_data)
            new_img = XLImage(img_stream)
            col_letter = chr(64 + col + 1)
            col_width_chars = ws.column_dimensions[col_letter].width or 12
            col_width_px = col_width_chars * 8
            row_height_pts = ws.row_dimensions[row].height or 15
            row_height_px = row_height_pts * 1.3333
            new_img.width = col_width_px * 0.95
            new_img.height = row_height_px * 0.95
            anchor = OneCellAnchor(
                _from=AnchorMarker(col=col, row=row-1, colOff=0, rowOff=0),
                ext=XDRPositiveSize2D(
                    cx=pixels_to_EMU(new_img.width),
                    cy=pixels_to_EMU(new_img.height)
                )
            )
            new_img.anchor = anchor
            ws.add_image(new_img)
            return True
        except Exception as e:
            print(f"            ⚠️ 插入失败 (行{row}列{chr(64+col+1)}): {e}")
            return False

    print("\n   📝 插入图片并填充数据...")

    img_idx1 = 0
    img_idx2 = 0
    total_inserted = 0
    data_filled_count = 0
    position_to_col = {'A': 0, 'B': 1, 'C': 2, 'D': 3}

    # 按顺序分配数据组给用户指定的 Fixture
    for group_idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            continue
        if group_idx >= len(fixture_data_groups):
            print(f"      ⚠️ 数据组不足，跳过 Fixture{fixture_num}")
            continue
        
        start_row, _ = target_ranges[fixture_num]
        group_data = fixture_data_groups[group_idx]  # 12行数据
        print(f"      Fixture{fixture_num}: 目标行 {start_row}-{start_row+11}")

        for row_offset in range(12):  # 0-11
            target_row = start_row + row_offset
            mapped_data = group_data[row_offset] if row_offset < len(group_data) else {}

            # ---- 插入 H 列 ----
            if img_idx1 < len(images1):
                if insert_image_to_cell(ws_target, target_row, 7, os.path.join(folder1, images1[img_idx1])):
                    total_inserted += 1
                img_idx1 += 1

            # ---- 插入 I 列 ----
            if img_idx2 < len(images2):
                if insert_image_to_cell(ws_target, target_row, 8, os.path.join(folder2, images2[img_idx2])):
                    total_inserted += 1
                img_idx2 += 1

            # ---- 填充 J-M 列 ----
            col_map = {'J': 10, 'K': 11, 'L': 12, 'M': 13}
            for col_key, val in mapped_data.items():
                if col_key in col_map:
                    safe_set_cell_value(ws_target, target_row, col_map[col_key], val)
                    data_filled_count += 1

        print(f"         ✅ Fixture{fixture_num} 数据填充完成")

    print(f"\n   ✅ X-Section 更新完成")
    print(f"   📊 共插入 {total_inserted} 张图片")
    print(f"   📊 共填充 {data_filled_count} 个数据值")
    print("=" * 60)

def update_visual_inspection_report(wb, data_folder_path, selected_fixtures):
    """
    更新 Visual Inspection Sheet
    1. 清空所有 Fixture 的 +2 行图片，+3 和 +4 行的数据
    2. 设置 +2 行的单元格尺寸：行高=96，栏宽=32.2
    3. 按用户指定 Fixture 顺序，从 IPQC Data 复制对应 Fixture 的 +2 行图片，+3 和 +4 行的数据
    4. 始终复制全部 10 列 (C-L)，不区分 MBO/PBO
    每个 Fixture 固定 5 行，连续排列
    """
    print("\n" + "=" * 60)
    print("📝 更新 Visual Inspection Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过更新")
        return

    print(f"   📌 用户选择的 Fixture: {selected_fixtures}")
    print(f"   📌 每个 Fixture 固定 5 行，连续排列（无空行分隔）")
    print(f"   📌 +2 行：图片 | +3 行：Spots number | +4 行：Pass/Fail")

    # ---- 目标行范围（每个Fixture 5行） ----
    target_ranges = {
        1: (2, 6), 2: (8, 12), 3: (14, 18), 4: (20, 24),
        5: (26, 30), 6: (32, 36), 7: (38, 42), 8: (44, 48),
    }

    # ---- 源数据行起始行（按顺序连续存放，每组5行） ----
    source_start_rows = [2, 8, 14, 20, 26, 32, 38, 44]

    # ---- 数据行偏移（相对于 Fixture 起始行） ----
    IMAGE_ROW_OFFSET = 2      # +2 行：图片
    DATA_ROW_OFFSETS = [3, 4]  # +3 行（Spots number），+4 行（Pass/Fail）

    # ---- 单元格尺寸（直接从Excel读取） ----
    ROW_HEIGHT = 96           # 磅
    COL_WIDTH = 32.2          # 字符宽度单位

    # ---- 获取目标 Sheet ----
    try:
        ws_target = wb["Visual Inspection "]
    except KeyError:
        print("   ⚠️ 未找到 'Visual Inspection ' Sheet，跳过更新")
        return

    # ============================================================
    # 清空数据与图片
    # ============================================================
    print("📝 清空所有 Fixture 的 +2 行图片，+3 和 +4 行的数据...")
    for fixture_num in range(1, 9):
        start_row, _ = target_ranges[fixture_num]
        
        # +2 行：清空图片
        img_row = start_row + IMAGE_ROW_OFFSET
        for col in range(3, 13):  # C-L（10列）
            ws_target.cell(row=img_row, column=col).value = None
        
        # +3 和 +4 行：清空数据
        for offset in DATA_ROW_OFFSETS:
            row = start_row + offset
            for col in range(3, 13):  # C-L（10列）
                ws_target.cell(row=row, column=col).value = None

    # 移除旧图片
    if ws_target._images:
        ws_target._images.clear()
        print("   ✅ 已移除所有旧图片")
    print("   ✅ 清空完成")

    # ============================================================
    # 设置 +2 行的单元格尺寸（行高=96，栏宽=32.2）
    # ============================================================
    print(f"📏 设置 +2 行的单元格尺寸：行高={ROW_HEIGHT}，栏宽={COL_WIDTH}...")
    
    for fixture_num in range(1, 9):
        start_row, _ = target_ranges[fixture_num]
        img_row = start_row + IMAGE_ROW_OFFSET
        ws_target.row_dimensions[img_row].height = ROW_HEIGHT

    for col in range(3, 13):
        ws_target.column_dimensions[get_column_letter(col)].width = COL_WIDTH
    print("   ✅ 尺寸设置完成")

    # ---- 定位 IPQC Data 文件 ----
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return

    ipqc_files = [
        f for f in os.listdir(ipqc_folder)
        if os.path.isfile(os.path.join(ipqc_folder, f))
        and not f.startswith('~$')
        and not f.startswith('.')
        and f.lower().endswith(('.xlsx', '.xlsm', '.xls'))
    ]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹中没有有效的 Excel 文件：{ipqc_folder}")
        return

    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")

    # ---- 读取 IPQC Data ----
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Visual Inspection " in wb_ipqc.sheetnames:
            ws_ipqc = wb_ipqc["Visual Inspection "]
            print("   ✅ 使用 IPQC Data 中的 'Visual Inspection ' Sheet")
        else:
            ws_ipqc = wb_ipqc.active
            print(f"   ⚠️ 使用第一个 Sheet：{ws_ipqc.title}")
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return

    # ============================================================
    # 提取图片（只针对 +2 行）
    # ============================================================
    print("📷 正在从 IPQC Data 提取图片...")
    image_map = {}  # key: (row, col), value: list of image data

    if hasattr(ws_ipqc, '_images') and ws_ipqc._images:
        print(f"   📷 在 IPQC Data 中找到 {len(ws_ipqc._images)} 张图片")
        for img in ws_ipqc._images:
            try:
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    from_cell = img.anchor._from
                    img_row = from_cell.row + 1
                    img_col = from_cell.col + 1
                    # 只收集 C-L 列（3-12），且行在 2-48 之间
                    if 2 <= img_row <= 48 and 3 <= img_col <= 12:
                        try:
                            img_data = img._data()
                            key = (img_row, img_col)
                            if key not in image_map:
                                image_map[key] = []
                            image_map[key].append(img_data)
                        except Exception:
                            pass
            except Exception:
                pass

    wb_ipqc.close()
    print(f"   ✅ 按位置归类了 {len(image_map)} 个单元格的图片")

    # ============================================================
    # 按顺序复制数据与图片（始终 10 列）
    # ============================================================
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from io import BytesIO

    print("\n   📝 按顺序复制数据与图片（C-L 列，10列）...")

    col_range = range(3, 13)  # C-L (10列)
    total_copied = 0

    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            print(f"      ⚠️ Fixture{fixture_num} 不在目标范围内，跳过")
            continue

        src_start = source_start_rows[idx]
        tgt_start, _ = target_ranges[fixture_num]

        print(f"      Fixture{fixture_num}: 源行 {src_start}-{src_start+4} → 目标行 {tgt_start}-{tgt_start+4}")

        # ---- 复制 +2 行图片 ----
        src_img_row = src_start + IMAGE_ROW_OFFSET
        tgt_img_row = tgt_start + IMAGE_ROW_OFFSET

        for col in col_range:
            key = (src_img_row, col)
            if key in image_map and image_map[key]:
                img_data = image_map[key][0]
                try:
                    img_stream = BytesIO(img_data)
                    new_img = XLImage(img_stream)

                    col_letter = get_column_letter(col)
                    # 使用已设置的列宽
                    col_width_chars = ws_target.column_dimensions[col_letter].width
                    if col_width_chars is None:
                        col_width_chars = COL_WIDTH
                    # 字符宽度 → 像素（近似）
                    col_width_px = col_width_chars * 7.5

                    row_height_pts = ws_target.row_dimensions[tgt_img_row].height
                    if row_height_pts is None:
                        row_height_pts = ROW_HEIGHT
                    # 磅 → 像素（近似）
                    row_height_px = row_height_pts * 1.3333

                    img_width_px = col_width_px * 0.95
                    img_height_px = row_height_px * 0.95

                    new_img.width = img_width_px
                    new_img.height = img_height_px

                    anchor = OneCellAnchor(
                        _from=AnchorMarker(col=col-1, row=tgt_img_row-1, colOff=0, rowOff=0),
                        ext=XDRPositiveSize2D(
                            cx=pixels_to_EMU(img_width_px),
                            cy=pixels_to_EMU(img_height_px)
                        )
                    )
                    new_img.anchor = anchor
                    ws_target.add_image(new_img)
                    total_copied += 1
                except Exception as e:
                    print(f"            ⚠️ 图片插入失败 (列{col_letter}行{tgt_img_row}): {e}")

        # ---- 复制 +3 和 +4 行数据 ----
        for offset in DATA_ROW_OFFSETS:
            src_row = src_start + offset
            tgt_row = tgt_start + offset
            for col in col_range:
                val = ws_ipqc.cell(row=src_row, column=col).value
                ws_target.cell(row=tgt_row, column=col, value=val)

    print(f"\n   ✅ Visual Inspection 更新完成，共复制 {total_copied} 张图片")
    print("=" * 60)

def update_outgassing_report(wb, data_folder_path, selected_fixtures):
    """
    更新 Outgassing Sheet
    1. 清除 E 列和 H 列的旧图片和数据
    2. 按用户指定 Fixture 顺序，从 IPQC Data 复制 E 列和 H 列的图片
    每个 Fixture 固定 3 行，连续排列：
    Fixture1: 行4-6, Fixture2: 行7-9, Fixture3: 行10-12, Fixture4: 行13-15,
    Fixture5: 行16-18, Fixture6: 行19-21, Fixture7: 行22-24, Fixture8: 行25-27
    """
    print("\n" + "=" * 60)
    print("📝 更新 Outgassing Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过更新")
        return

    print(f"   📌 用户选择的 Fixture: {selected_fixtures}")
    print(f"   📌 每个 Fixture 固定 3 行，连续排列（无空行分隔）")
    print(f"   📌 只处理 E 列（VCM backside image）和 H 列（IRCF image）的图片")

    # ---- 目标行范围（按Fixture编号，固定3行，连续排列） ----
    target_ranges = {
        1: (4, 6), 2: (7, 9), 3: (10, 12), 4: (13, 15),
        5: (16, 18), 6: (19, 21), 7: (22, 24), 8: (25, 27),
    }

    # ---- 源数据行起始行（按顺序连续存放，每组3行） ----
    source_start_rows = [4, 7, 10, 13, 16, 19, 22, 25]
    rows_per_fixture = 3  # 固定3行

    # ---- 获取目标 Sheet ----
    try:
        ws_target = wb["Outgassing"]
    except KeyError:
        print("   ⚠️ 未找到 'Outgassing' Sheet，跳过更新")
        return

    # ---- 清空 E 列和 H 列的数据，并移除所有旧图片 ----
    print("📝 清空 E 列和 H 列的数据并移除旧图片...")
    for row in range(4, 28):  # 最大行27
        ws_target.cell(row=row, column=5).value = None
        ws_target.cell(row=row, column=8).value = None

    if ws_target._images:
        ws_target._images.clear()
        print("   ✅ 已移除所有旧图片")
    print("   ✅ 清空完成")

    # ---- 定位 IPQC Data 文件 ----
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return

    ipqc_files = [
        f for f in os.listdir(ipqc_folder)
        if os.path.isfile(os.path.join(ipqc_folder, f))
        and not f.startswith('~$')
        and not f.startswith('.')
        and f.lower().endswith(('.xlsx', '.xlsm', '.xls'))
    ]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹中没有有效的 Excel 文件：{ipqc_folder}")
        return

    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")

    # ---- 读取 IPQC Data ----
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Outgassing" in wb_ipqc.sheetnames:
            ws_ipqc = wb_ipqc["Outgassing"]
            print("   ✅ 使用 IPQC Data 中的 'Outgassing' Sheet")
        else:
            ws_ipqc = wb_ipqc.active
            print(f"   ⚠️ 使用第一个 Sheet：{ws_ipqc.title}")
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return

    # ============================================================
    # 提取 E 列和 H 列的图片
    # ============================================================
    print("📷 正在从 IPQC Data 提取图片...")
    image_map = {}  # key: (row, col), value: list of image data

    if hasattr(ws_ipqc, '_images') and ws_ipqc._images:
        print(f"   📷 在 IPQC Data 中找到 {len(ws_ipqc._images)} 张图片")
        for img in ws_ipqc._images:
            try:
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    from_cell = img.anchor._from
                    img_row = from_cell.row + 1
                    img_col = from_cell.col + 1
                    # 只收集 E 列（5）和 H 列（8）的图片，且行在 4-27 之间
                    if 4 <= img_row <= 27 and img_col in [5, 8]:
                        try:
                            img_data = img._data()
                            key = (img_row, img_col)
                            if key not in image_map:
                                image_map[key] = []
                            image_map[key].append(img_data)
                        except Exception:
                            pass
            except Exception:
                pass

    wb_ipqc.close()
    print(f"   ✅ 按位置归类了 {len(image_map)} 个单元格的图片")

    if not image_map:
        print("   ⚠️ 没有找到可用的图片数据")
        return

    # ---- 按顺序复制图片 ----
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from io import BytesIO

    print("\n   📝 按顺序复制图片...")

    total_copied = 0

    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            print(f"      ⚠️ Fixture{fixture_num} 不在目标范围内，跳过")
            continue

        src_start = source_start_rows[idx]
        tgt_start, _ = target_ranges[fixture_num]
        rows_to_copy = rows_per_fixture

        print(f"      Fixture{fixture_num}: 源行 {src_start}-{src_start+rows_to_copy-1} → 目标行 {tgt_start}-{tgt_start+rows_to_copy-1}")

        for row_offset in range(rows_to_copy):
            src_row = src_start + row_offset
            tgt_row = tgt_start + row_offset

            for col in [5, 8]:  # E 和 H
                key = (src_row, col)
                if key not in image_map or not image_map[key]:
                    continue

                img_data = image_map[key][0]
                try:
                    img_stream = BytesIO(img_data)
                    new_img = XLImage(img_stream)

                    col_letter = get_column_letter(col)
                    col_width_chars = ws_target.column_dimensions[col_letter].width
                    if col_width_chars is None:
                        col_width_chars = 12
                    col_width_px = col_width_chars * 8

                    row_height_pts = ws_target.row_dimensions[tgt_row].height
                    if row_height_pts is None:
                        row_height_pts = 15
                    row_height_px = row_height_pts * 1.3333

                    # 图片尺寸设为单元格的 95%（留边）
                    img_width_px = col_width_px * 0.95
                    img_height_px = row_height_px * 0.95

                    new_img.width = img_width_px
                    new_img.height = img_height_px

                    anchor = OneCellAnchor(
                        _from=AnchorMarker(col=col-1, row=tgt_row-1, colOff=0, rowOff=0),
                        ext=XDRPositiveSize2D(
                            cx=pixels_to_EMU(img_width_px),
                            cy=pixels_to_EMU(img_height_px)
                        )
                    )
                    new_img.anchor = anchor
                    ws_target.add_image(new_img)
                    total_copied += 1
                except Exception as e:
                    print(f"            ⚠️ 插入失败 (列{col_letter}行{tgt_row}): {e}")

    print(f"\n   ✅ Outgassing 更新完成，共复制 {total_copied} 张图片")
    print("=" * 60)

def update_bond_strength_report(wb, data_folder_path, selected_fixtures, 报告类型):
    """
    更新 Bond Strength - Push Test Sheet
    1. 复制 E 列数值数据（按顺序取源数据）
    2. 补全 H/I/J/K 列空值（从 Sheet 第10行复制）
    3. 复制图片（按列顺序 E→F→G→I→J→K）
       - E 列：宽100%，高90%（底部留空显示数值）
       - F/G 列：宽100%，高100%
       - I/J/K 列：第一张图片宽50%，高90%；第二张图片宽50%，高90%，左右并排（第二张右移第一张图片宽度）
    4. 复制 I/J/K 文本数据
    """
    print("\n" + "=" * 60)
    print("📝 更新 Bond Strength - Push Test Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过更新")
        return

    print(f"   📌 用户选择的 Fixture: {selected_fixtures}")

    rows_per_fixture = 7 if 报告类型 == "MBO" else 10
    print(f"   📌 报告类型：{报告类型}，每个 Fixture {rows_per_fixture} 行数据")

    # ---- 目标行范围（按Fixture编号） ----
    pbo_target = {
        1: (10, 19), 2: (20, 29), 3: (30, 39), 4: (40, 49),
        5: (50, 59), 6: (60, 69), 7: (70, 79), 8: (80, 89),
    }
    mbo_target = {
        1: (10, 16), 2: (20, 26), 3: (30, 36), 4: (40, 46),
        5: (50, 56), 6: (60, 66), 7: (70, 76), 8: (80, 86),
    }
    target_ranges = mbo_target if 报告类型 == "MBO" else pbo_target

    # ---- 源数据行起始行（按顺序连续存放） ----
    source_start_rows = [10, 20, 30, 40, 50, 60, 70, 80]

    # ---- 获取目标 Sheet ----
    try:
        ws_target = wb["Bond Strength - Push Test"]
    except KeyError:
        print("   ⚠️ 未找到 'Bond Strength - Push Test' Sheet，跳过更新")
        return

    # ---- 清空 ----
    print("📝 清空数据区域并移除旧图片...")
    for row in range(10, 90):
        for col in [5, 6, 7, 8, 9, 10, 11]:
            ws_target.cell(row=row, column=col).value = None

    if ws_target._images:
        ws_target._images.clear()
        print("   ✅ 已移除所有旧图片")
    print("   ✅ 清空完成")

    # ---- MBO 额外清空 ----
    if 报告类型 == "MBO":
        print("📝 MBO 报告：清空每个 Fixture 的第 8-10 行...")
        mbo_extra_rows = {
            1: (17, 19), 2: (27, 29), 3: (37, 39), 4: (47, 49),
            5: (57, 59), 6: (67, 69), 7: (77, 79), 8: (87, 89),
        }
        for fixture_num in selected_fixtures:
            if fixture_num in mbo_extra_rows:
                start_row, end_row = mbo_extra_rows[fixture_num]
                for row in range(start_row, end_row + 1):
                    for col in [5, 6, 7, 8, 9, 10, 11]:
                        ws_target.cell(row=row, column=col).value = None
                print(f"      Fixture{fixture_num}: 清空行 {start_row}-{end_row}")
        print("   ✅ MBO 清空完成")

    # ---- 定位 IPQC Data 文件 ----
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return

    ipqc_files = [
        f for f in os.listdir(ipqc_folder)
        if os.path.isfile(os.path.join(ipqc_folder, f))
        and not f.startswith('~$')
        and not f.startswith('.')
        and f.lower().endswith(('.xlsx', '.xlsm', '.xls'))
    ]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹中没有有效的 Excel 文件：{ipqc_folder}")
        return

    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")

    # ---- 读取 IPQC Data ----
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Bond Strength - Push Test" in wb_ipqc.sheetnames:
            ws_ipqc = wb_ipqc["Bond Strength - Push Test"]
        else:
            ws_ipqc = wb_ipqc.active
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return

    # ============================================================
    # 1. 先复制 E 列数值数据（非图片）
    # ============================================================
    print("\n   📝 复制 E 列数值数据...")
    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            print(f"      ⚠️ Fixture{fixture_num} 不在目标范围内，跳过")
            continue
        src_start = source_start_rows[idx]
        tgt_start, _ = target_ranges[fixture_num]
        rows_to_copy = rows_per_fixture
        for row_offset in range(rows_to_copy):
            src_row = src_start + row_offset
            tgt_row = tgt_start + row_offset
            val = ws_ipqc.cell(row=src_row, column=5).value
            ws_target.cell(row=tgt_row, column=5, value=val)

    # ============================================================
    # 2. 补全 H/I/J/K 列空值（从 Sheet 第10行复制）
    #    第10行是固定行号，即第一个Fixture的起始行
    # ============================================================
    print("\n   📝 补全 H/I/J/K 列空值（从第10行复制默认值）...")
    # 读取第10行的 H/I/J/K 默认值
    default_h = ws_ipqc.cell(row=10, column=8).value   # H
    default_i = ws_ipqc.cell(row=10, column=9).value   # I
    default_j = ws_ipqc.cell(row=10, column=10).value  # J
    default_k = ws_ipqc.cell(row=10, column=11).value  # K

    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            continue
        tgt_start, _ = target_ranges[fixture_num]
        rows_to_copy = rows_per_fixture
        for row_offset in range(rows_to_copy):
            tgt_row = tgt_start + row_offset
            # 检查 H 列
            if ws_target.cell(row=tgt_row, column=8).value is None:
                ws_target.cell(row=tgt_row, column=8, value=default_h)
            # 检查 I 列
            if ws_target.cell(row=tgt_row, column=9).value is None:
                ws_target.cell(row=tgt_row, column=9, value=default_i)
            # 检查 J 列
            if ws_target.cell(row=tgt_row, column=10).value is None:
                ws_target.cell(row=tgt_row, column=10, value=default_j)
            # 检查 K 列
            if ws_target.cell(row=tgt_row, column=11).value is None:
                ws_target.cell(row=tgt_row, column=11, value=default_k)

    # ============================================================
    # 3. 提取图片并按列顺序复制（调整大小和偏移）
    # ============================================================
    print("📷 正在从 IPQC Data 提取图片...")
    image_map = {}  # key: (row, col), value: list of image data
    if hasattr(ws_ipqc, '_images') and ws_ipqc._images:
        print(f"   📷 在 IPQC Data 中找到 {len(ws_ipqc._images)} 张图片")
        for img in ws_ipqc._images:
            try:
                if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                    from_cell = img.anchor._from
                    img_row = from_cell.row + 1
                    img_col = from_cell.col + 1
                    if 10 <= img_row <= 89 and 5 <= img_col <= 15:
                        try:
                            img_data = img._data()
                            key = (img_row, img_col)
                            if key not in image_map:
                                image_map[key] = []
                            image_map[key].append(img_data)
                        except Exception:
                            pass
            except Exception:
                pass

    wb_ipqc.close()
    print(f"   ✅ 按位置归类了 {len(image_map)} 个单元格的图片")

    if not image_map:
        print("   ⚠️ 没有找到可用的图片数据")
        return

    # ---- 按列顺序复制图片 ----
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    from io import BytesIO

    print("\n   📝 按列顺序复制图片...")

    # 列配置：(列号, 列名, 宽度比例, 高度比例)
    # I/J/K 列：第一张图片宽度50%，第二张图片宽度50%，高度90%
    col_configs = [
        (5, "E", 1.0, 0.9),
        (6, "F", 1.0, 1.0),
        (7, "G", 1.0, 1.0),
        (9, "I", 0.5, 0.9),
        (10, "J", 0.5, 0.9),
        (11, "K", 0.5, 0.9),
    ]

    total_copied = 0

    for col, label, width_ratio, height_ratio in col_configs:
        print(f"      📌 处理列 {label}...")
        col_copied = 0

        col_letter = get_column_letter(col)
        col_width_chars = ws_target.column_dimensions[col_letter].width
        if col_width_chars is None:
            col_width_chars = 12
        col_width_px = col_width_chars * 8

        for idx, fixture_num in enumerate(selected_fixtures):
            if fixture_num not in target_ranges:
                continue

            src_start = source_start_rows[idx]
            tgt_start, _ = target_ranges[fixture_num]
            rows_to_copy = rows_per_fixture

            for row_offset in range(rows_to_copy):
                tgt_row = tgt_start + row_offset
                src_row = src_start + row_offset

                key = (src_row, col)
                if key not in image_map or not image_map[key]:
                    continue

                row_height_pts = ws_target.row_dimensions[tgt_row].height
                if row_height_pts is None:
                    row_height_pts = 15
                row_height_px = row_height_pts * 1.3333

                # ---- I/J/K 列双图片 ----
                if col in [9, 10, 11]:
                    # 第一张图片：宽度 50%，高度 90%
                    img_data1 = image_map[key][0]
                    try:
                        img_stream1 = BytesIO(img_data1)
                        new_img1 = XLImage(img_stream1)
                        img1_width_px = col_width_px * 0.50
                        img1_height_px = row_height_px * 0.90
                        new_img1.width = img1_width_px
                        new_img1.height = img1_height_px

                        anchor1 = OneCellAnchor(
                            _from=AnchorMarker(col=col-1, row=tgt_row-1, colOff=0, rowOff=0),
                            ext=XDRPositiveSize2D(
                                cx=pixels_to_EMU(img1_width_px),
                                cy=pixels_to_EMU(img1_height_px)
                            )
                        )
                        new_img1.anchor = anchor1
                        ws_target.add_image(new_img1)
                        col_copied += 1
                        total_copied += 1
                    except Exception as e:
                        print(f"            ⚠️ 插入失败 (列{label}行{tgt_row} 左图): {e}")

                    # 第二张图片：宽度 50%，高度 90%，右移第一张图片宽度
                    if len(image_map[key]) > 1:
                        img_data2 = image_map[key][1]
                        try:
                            img_stream2 = BytesIO(img_data2)
                            new_img2 = XLImage(img_stream2)
                            img2_width_px = col_width_px * 0.50
                            img2_height_px = row_height_px * 0.90
                            new_img2.width = img2_width_px
                            new_img2.height = img2_height_px

                            offset_emu = pixels_to_EMU(img1_width_px)
                            anchor2 = OneCellAnchor(
                                _from=AnchorMarker(col=col-1, row=tgt_row-1, colOff=offset_emu, rowOff=0),
                                ext=XDRPositiveSize2D(
                                    cx=pixels_to_EMU(img2_width_px),
                                    cy=pixels_to_EMU(img2_height_px)
                                )
                            )
                            new_img2.anchor = anchor2
                            ws_target.add_image(new_img2)
                            col_copied += 1
                            total_copied += 1
                        except Exception as e:
                            print(f"            ⚠️ 插入失败 (列{label}行{tgt_row} 右图): {e}")
                else:
                    # ---- E/F/G 列单图片 ----
                    img_data = image_map[key][0]
                    try:
                        img_stream = BytesIO(img_data)
                        new_img = XLImage(img_stream)
                        img_width_px = col_width_px * width_ratio
                        img_height_px = row_height_px * height_ratio
                        new_img.width = img_width_px
                        new_img.height = img_height_px
                        cell_ref = f"{col_letter}{tgt_row}"
                        ws_target.add_image(new_img, cell_ref)
                        col_copied += 1
                        total_copied += 1
                    except Exception as e:
                        print(f"            ⚠️ 插入失败 (列{label}行{tgt_row}): {e}")

        print(f"         列 {label} 共复制 {col_copied} 张图片")

    # ============================================================
    # 4. 复制 I/J/K 文本数据（按顺序取源数据）
    # ============================================================
    print("\n   📝 复制文本数据 (I, J, K 列)...")
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        ws_ipqc = wb_ipqc["Bond Strength - Push Test"] if "Bond Strength - Push Test" in wb_ipqc.sheetnames else wb_ipqc.active
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return

    text_copied = 0
    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in target_ranges:
            continue
        src_start = source_start_rows[idx]
        tgt_start, _ = target_ranges[fixture_num]
        rows_to_copy = rows_per_fixture
        for row_offset in range(rows_to_copy):
            src_row = src_start + row_offset
            tgt_row = tgt_start + row_offset
            ws_target.cell(row=tgt_row, column=9, value=ws_ipqc.cell(row=src_row, column=9).value)
            ws_target.cell(row=tgt_row, column=10, value=ws_ipqc.cell(row=src_row, column=10).value)
            ws_target.cell(row=tgt_row, column=11, value=ws_ipqc.cell(row=src_row, column=11).value)
            text_copied += 3
    wb_ipqc.close()

    print("\n   ✅ Bond Strength - Push Test 更新完成")
    print(f"   📊 共复制了 {total_copied} 张图片，{text_copied} 个文本数据，以及 E 列数值")
    print("=" * 60)


def update_weld_spot_diameter_report(wb, data_folder_path, selected_fixtures):
    """
    更新 Weld Spot Diameter Sheet
    从 IPQC Data 文件夹中的文件读取数据
    Fixture 横向排列（C列到J列），IPQC Data 与报告格式完全一致
    
    数据清洗逻辑：
    1. 检查第5-39行每个 Fixture 的数据
    2. 如果数值不在 0.45~0.50 范围内，标记为异常
    3. 从第40-54行（备用数据）中挑选合格数值进行替换
    """
    print("\n" + "=" * 60)
    print("📝 更新 Weld Spot Diameter Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过更新")
        return
    
    # 获取报告中的 Sheet
    try:
        ws_target = wb["Weld Spot Diameter"]
    except KeyError:
        print("   ⚠️ 未找到 'Weld Spot Diameter' Sheet，跳过更新")
        return
    
    # 清空数据区域（C5:J39）
    print("📝 清空 Weld Spot Diameter 数据区域...")
    for row in range(5, 40):
        for col in range(3, 11):  # C=3, J=10
            ws_target.cell(row=row, column=col).value = None
    print("   ✅ 清空完成")
    
    # 读取 IPQC Data 文件夹
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return
    
    ipqc_files = [f for f in os.listdir(ipqc_folder) if os.path.isfile(os.path.join(ipqc_folder, f))]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹为空：{ipqc_folder}")
        return
    
    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")
    
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Weld Spot Diameter" in wb_ipqc.sheetnames:
            ws_ipqc = wb_ipqc["Weld Spot Diameter"]
            print("   ✅ 使用 IPQC Data 中的 'Weld Spot Diameter' Sheet")
        else:
            ws_ipqc = wb_ipqc.active
            print(f"   ⚠️ IPQC Data 文件中没有 'Weld Spot Diameter' Sheet，使用第一个 Sheet：{ws_ipqc.title}")
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return
    
    # ============================================================
    # Fixture 编号 → 列索引映射
    # C=3, D=4, E=5, F=6, G=7, H=8, I=9, J=10
    # ============================================================
    fixture_column_mapping = {
        1: 3, 2: 4, 3: 5, 4: 6,
        5: 7, 6: 8, 7: 9, 8: 10,
    }
    
    # 合格范围
    SPEC_LOW = 0.45
    SPEC_HIGH = 0.50
    
    print("\n   📝 Weld Spot Diameter 数据清洗与映射关系：")
    
    # 先收集所有需要替换的异常值位置
    abnormal_positions = {}  # {fixture_num: [(row, current_value), ...]}
    
    for fixture_num in selected_fixtures:
        if fixture_num not in fixture_column_mapping:
            print(f"      ⚠️ 无效的 Fixture 编号：{fixture_num}，跳过")
            continue
        
        col = fixture_column_mapping[fixture_num]
        col_letter = get_column_letter(col)
        abnormal_list = []
        
        # 检查第5-39行的数据
        for row in range(5, 40):
            val = ws_ipqc.cell(row=row, column=col).value
            # 检查数值是否在合格范围内
            if val is not None:
                try:
                    num_val = float(val)
                    if num_val < SPEC_LOW or num_val > SPEC_HIGH:
                        abnormal_list.append((row, num_val))
                except (ValueError, TypeError):
                    # 如果无法转换为数字，也视为异常
                    abnormal_list.append((row, val))
        
        if abnormal_list:
            abnormal_positions[fixture_num] = abnormal_list
            print(f"      Fixture{fixture_num} (列{col_letter}): 发现 {len(abnormal_list)} 个异常值")
        else:
            print(f"      Fixture{fixture_num} (列{col_letter}): ✅ 全部合格")
    
    # ============================================================
    # 从备用数据（第40-54行）中挑选合格数值进行替换
    # ============================================================
    if abnormal_positions:
        print("\n   📝 开始从备用数据中挑选合格数值进行替换...")
        
        for fixture_num, positions in abnormal_positions.items():
            col = fixture_column_mapping[fixture_num]
            col_letter = get_column_letter(col)
            
            # 收集备用数据中所有合格的数值（第40-54行）
            backup_valid_values = []
            for row in range(40, 55):
                val = ws_ipqc.cell(row=row, column=col).value
                if val is not None:
                    try:
                        num_val = float(val)
                        if SPEC_LOW <= num_val <= SPEC_HIGH:
                            backup_valid_values.append(num_val)
                    except (ValueError, TypeError):
                        pass
            
            print(f"      Fixture{fixture_num} (列{col_letter}): 备用数据中找到 {len(backup_valid_values)} 个合格数值")
            
            if not backup_valid_values:
                print(f"         ⚠️ 备用数据中没有找到合格数值，无法替换 Fixture{fixture_num} 的异常值")
                continue
            
            # 逐个替换异常值
            replaced_count = 0
            for row, current_val in positions:
                if backup_valid_values:
                    # 取第一个备用合格值
                    new_val = backup_valid_values.pop(0)
                    ws_ipqc.cell(row=row, column=col, value=new_val)
                    replaced_count += 1
                    print(f"         行{row}: {current_val} → {new_val} ✓")
                else:
                    print(f"         行{row}: {current_val} → 备用合格值已用完，无法替换")
            
            print(f"      Fixture{fixture_num}: 成功替换 {replaced_count} 个异常值")
    
    # ============================================================
    # 将清洗后的数据复制到报告
    # ============================================================
    print("\n   📝 将清洗后的数据复制到报告...")
    
    for fixture_num in selected_fixtures:
        if fixture_num not in fixture_column_mapping:
            continue
        
        col = fixture_column_mapping[fixture_num]
        col_letter = get_column_letter(col)
        
        print(f"      IPQC 列 {col_letter} → 报告列 {col_letter} (Fixture{fixture_num})")
        
        # 复制数据：从第5行到第39行
        for row in range(5, 40):
            val = ws_ipqc.cell(row=row, column=col).value
            ws_target.cell(row=row, column=col, value=val)
    
    wb_ipqc.close()
    print("\n   ✅ Weld Spot Diameter 更新完成")
    print("=" * 60)

def get_fixture_selection():
    """
    使用 GUI 弹窗获取用户输入的 Fixture 编号
    返回：选中的 Fixture 列表（按顺序）
    """
    import tkinter as tk
    from tkinter import simpledialog

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # 构建提示信息
    prompt = (
        "请输入需要更新的 Fixture 编号（用空格隔开，如：1 2 7 8）\n\n"
        "1: Fixture1  2: Fixture2  3: Fixture3  4: Fixture4\n"
        "5: Fixture5  6: Fixture6  7: Fixture7  8: Fixture8\n\n"
        "⚠️ 输入顺序必须与 IPQC Data 中的存储顺序一致"
    )

    input_str = simpledialog.askstring(
        title="选择 Fixture",
        prompt=prompt,
        parent=root
    )
    root.destroy()

    if not input_str:
        print("   ⚠️ 未选择任何 Fixture")
        return []

    try:
        selected = [int(x.strip()) for x in input_str.split() if x.strip().isdigit()]
        # 去重但保留顺序
        seen = set()
        result = []
        for f in selected:
            if f not in seen:
                seen.add(f)
                result.append(f)

        invalid = [f for f in result if f not in range(1, 9)]
        if invalid:
            print(f"   ⚠️ 无效的 Fixture 编号：{invalid}，请使用 1-8 的数字")
            return []

        if len(result) > 8:
            print(f"   ⚠️ 最多支持 8 个 Fixture，将只取前 8 个")
            result = result[:8]

        print(f"   ✅ 选择的 Fixture（按顺序）: {result}")
        return result
    except:
        print("   ⚠️ 输入格式错误")
        return []


def update_dcr_report(wb, data_folder_path, selected_fixtures):
    """
    更新 DCR Sheet
    从 IPQC Data 文件夹中的文件读取 DCR 数据，直接复制粘贴到报告
    """
    print("\n" + "=" * 60)
    print("📝 更新 DCR Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过 DCR 更新")
        return
    
    # 获取报告中的 DCR Sheet
    try:
        ws_dcr = wb["DCR"]
    except KeyError:
        print("   ⚠️ 未找到 'DCR' Sheet，跳过更新")
        return
    
    # 清空 DCR 指定区域的内容（保留格式）
    print("📝 清空 DCR 指定区域...")
    
    # 清空 D4:E43
    for row in range(4, 44):
        for col in range(4, 6):  # D=4, E=5
            ws_dcr.cell(row=row, column=col).value = None
    
    # 清空 D53:E92
    for row in range(53, 93):
        for col in range(4, 6):  # D=4, E=5
            ws_dcr.cell(row=row, column=col).value = None
    
    print("   ✅ 清空完成")
    
    # 读取 IPQC Data 文件夹
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return
    
    ipqc_files = [f for f in os.listdir(ipqc_folder) if os.path.isfile(os.path.join(ipqc_folder, f))]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹为空：{ipqc_folder}")
        return
    
    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")
    
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        
        if "DCR" in wb_ipqc.sheetnames:
            ws_ipqc_dcr = wb_ipqc["DCR"]
            print("   ✅ 使用 IPQC Data 中的 'DCR' Sheet")
        else:
            ws_ipqc_dcr = wb_ipqc.active
            print(f"   ⚠️ IPQC Data 文件中没有 'DCR' Sheet，使用第一个 Sheet：{ws_ipqc_dcr.title}")
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return
    
    # IPQC Data 中的 8 个 DCR 数据区域（D列到E列）
    ipqc_dcr_ranges = [
        (4, 13), (14, 23), (24, 33), (34, 43),
        (53, 62), (63, 72), (73, 82), (83, 92),
    ]
    
    # 报告中 DCR Fixture 对应的行范围（D列到E列）
    dcr_fixture_row_mapping = {
        1: (4, 13), 2: (14, 23), 3: (24, 33), 4: (34, 43),
        5: (53, 62), 6: (63, 72), 7: (73, 82), 8: (83, 92),
    }
    
    print("\n   📝 DCR 数据映射关系：")
    
    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in dcr_fixture_row_mapping:
            print(f"      ⚠️ 无效的 Fixture 编号：{fixture_num}，跳过")
            continue
        
        if idx >= len(ipqc_dcr_ranges):
            print(f"      ⚠️ 第 {idx+1} 组数据不存在，跳过 Fixture{fixture_num}")
            continue
        
        src_start, src_end = ipqc_dcr_ranges[idx]
        tgt_start, tgt_end = dcr_fixture_row_mapping[fixture_num]
        
        print(f"      IPQC D{src_start}:E{src_end} → Fixture{fixture_num} D{tgt_start}:E{tgt_end}")
        
        for row_offset in range(10):
            src_row = src_start + row_offset
            tgt_row = tgt_start + row_offset
            for col_offset in range(2):
                src_col = 4 + col_offset
                tgt_col = 4 + col_offset
                val = ws_ipqc_dcr.cell(row=src_row, column=src_col).value
                ws_dcr.cell(row=tgt_row, column=tgt_col, value=val)
    
    wb_ipqc.close()
    print("\n   ✅ DCR 更新完成")
    print("=" * 60)


def update_raw_data_report(wb, data_folder_path, selected_fixtures):
    """
    更新 Raw Data Sheet
    从 IPQC Data 文件夹中的文件读取数据，直接复制粘贴到报告
    """
    print("\n" + "=" * 60)
    print("📝 更新 Raw Data Sheet...")
    print("=" * 60)
    
    if not selected_fixtures:
        print("   ⚠️ 未选择任何 Fixture，跳过 Raw Data 更新")
        return
    
    # 获取 Raw Data Sheet
    try:
        ws_raw = wb["Raw Data"]
    except KeyError:
        print("   ⚠️ 未找到 'Raw Data' Sheet，跳过更新")
        return
    
    # 清空指定区域
    print("📝 清空 Raw Data 指定区域...")
    for row in range(10, 50):
        for col in range(5, 18):
            ws_raw.cell(row=row, column=col).value = None
    for row in range(64, 104):
        for col in range(5, 18):
            ws_raw.cell(row=row, column=col).value = None
    print("   ✅ 清空完成")
    
    # 读取 IPQC Data 文件夹
    ipqc_folder = os.path.join(data_folder_path, "IPQC Data")
    if not os.path.exists(ipqc_folder):
        print(f"   ⚠️ IPQC Data 文件夹不存在：{ipqc_folder}")
        return
    
    ipqc_files = [f for f in os.listdir(ipqc_folder) if os.path.isfile(os.path.join(ipqc_folder, f))]
    if not ipqc_files:
        print(f"   ⚠️ IPQC Data 文件夹为空：{ipqc_folder}")
        return
    
    ipqc_file = os.path.join(ipqc_folder, ipqc_files[0])
    print(f"   📄 读取 IPQC 数据文件：{ipqc_files[0]}")
    
    try:
        wb_ipqc = openpyxl.load_workbook(ipqc_file, data_only=True)
        if "Raw Data" in wb_ipqc.sheetnames:
            ws_ipqc = wb_ipqc["Raw Data"]
            print("   ✅ 使用 IPQC Data 中的 'Raw Data' Sheet")
        else:
            ws_ipqc = wb_ipqc.active
            print(f"   ⚠️ IPQC Data 文件中没有 'Raw Data' Sheet，使用第一个 Sheet：{ws_ipqc.title}")
    except Exception as e:
        print(f"   ❌ 读取 IPQC 文件失败：{e}")
        return
    
    # IPQC Data 中的 8 个数据区域
    ipqc_ranges = [
        (10, 19), (20, 29), (30, 39), (40, 49),
        (64, 73), (74, 83), (84, 93), (94, 103),
    ]
    
    fixture_row_mapping = {
        1: (10, 19), 2: (20, 29), 3: (30, 39), 4: (40, 49),
        5: (64, 73), 6: (74, 83), 7: (84, 93), 8: (94, 103),
    }
    
    print("\n   📝 Raw Data 映射关系：")
    
    for idx, fixture_num in enumerate(selected_fixtures):
        if fixture_num not in fixture_row_mapping:
            print(f"      ⚠️ 无效的 Fixture 编号：{fixture_num}，跳过")
            continue
        
        if idx >= len(ipqc_ranges):
            print(f"      ⚠️ 第 {idx+1} 组数据不存在，跳过 Fixture{fixture_num}")
            continue
        
        src_start, src_end = ipqc_ranges[idx]
        tgt_start, tgt_end = fixture_row_mapping[fixture_num]
        
        print(f"      IPQC E{src_start}:Q{src_end} → Fixture{fixture_num} E{tgt_start}:Q{tgt_end}")
        
        for row_offset in range(10):
            src_row = src_start + row_offset
            tgt_row = tgt_start + row_offset
            for col_offset in range(13):
                src_col = 5 + col_offset
                tgt_col = 5 + col_offset
                val = ws_ipqc.cell(row=src_row, column=src_col).value
                ws_raw.cell(row=tgt_row, column=tgt_col, value=val)
    
    wb_ipqc.close()
    print("\n   ✅ Raw Data 更新完成")
    print("=" * 60)


def update_summary_report(template_path, data_folder_path, output_folder_path=None):
    """
    更新Summary报告的主要函数
    """
    # 1. 获取当天日期
    today = datetime.now()
    today_str = today.strftime("%Y%m%d")
    today_slash = today.strftime("%Y/%m/%d")
    
    # 2. 解析文件夹名称
    folder_name = os.path.basename(data_folder_path)
    parsed = parse_folder_name(folder_name)
    if not parsed:
        return
    
    专案 = parsed["专案"]
    阶段 = parsed["阶段"]
    Config = parsed["Config"]
    机台号 = parsed["机台号"]
    报告类型 = parsed["报告类型"]
    
    print("=" * 60)
    print(f"📁 数据文件夹：{folder_name}")
    print(f"   📌 专案：{专案}")
    print(f"   📌 阶段：{阶段}")
    print(f"   📌 Config：{Config}")
    print(f"   📌 机台号：{机台号}")
    print(f"   📌 报告类型：{报告类型}")
    print("=" * 60)
    
    # 3. 生成新报告名称
    report_name = f"{专案}_ABU_{Config}_{报告类型}_Chassis_Attach(Welding)_{机台号}_{today_str}_Rev.0"
    print(f"📄 新报告名称：{report_name}")
    
    # 4. 复制模板文件
    if not os.path.exists(template_path):
        print(f"❌ 错误：找不到模板文件 {template_path}")
        return
    
    if output_folder_path is None:
        output_folder_path = data_folder_path
    
    output_path = os.path.join(output_folder_path, f"{report_name}.xlsx")
    
    shutil.copy2(template_path, output_path)
    print(f"✅ 已复制模板到：{output_path}")
    
    # 5. 打开Excel文件进行编辑
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # ============================================================
    # 【第一步】获取用户输入的 Fixture（只输入一次）
    # ============================================================
    selected_fixtures = get_fixture_selection()
    
    if not selected_fixtures:
        print("⚠️ 未选择任何 Fixture，将跳过 Raw Data 和 DCR 更新")
    
    # ============================================================
    # 【第二步】3.1 + 3.2：历史数据归档（必须先执行！）
    # ============================================================
    print("=" * 60)
    print("📝 【历史数据归档】开始执行...")
    print("=" * 60)
    
    # 3.1 将K25到BC185整个区域剪切到L25到BD185
    print("📝 步骤3.1：剪切 K25:BC185 → L25:BD185...")
    print("   ⚠️ 注意：此操作会取消合并单元格")
    
    data_region = []
    for row in range(25, 186):
        row_data = []
        for col in range(11, 56):
            row_data.append(ws.cell(row=row, column=col).value)
        data_region.append(row_data)
    
    for row in range(25, 186):
        for col in range(11, 56):
            ws.cell(row=row, column=col, value=None)
    
    for row_offset, row_data in enumerate(data_region):
        row = 25 + row_offset
        for col_offset, value in enumerate(row_data):
            col = 12 + col_offset
            ws.cell(row=row, column=col, value=value)
    
    print("   ✅ 剪切完成（历史数据已移至 L25:BD185）")
    
    # 3.2 复制L25到L185的格式至K25到K185
    print("📝 步骤3.2：复制格式 L25:L185 → K25:K185...")
    for row in range(25, 186):
        source_cell = ws.cell(row=row, column=12)
        target_cell = ws.cell(row=row, column=11)
        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.alignment = source_cell.alignment.copy()
    
    print("   ✅ 格式复制完成（K列已准备就绪）")
    print("=" * 60)
    print("✅ 【历史数据归档】完成，开始填充新数据...")
    print("=" * 60)
    
    # ============================================================
    # 【第三步】3.4 复制I列到K列（26-155行，159-185行）
    # ============================================================
    print("📝 步骤3.4：复制 I列 → K列...")
    for row in list(range(26, 156)) + list(range(160, 186)):
        source_value = ws.cell(row=row, column=9).value
        ws.cell(row=row, column=11, value=source_value)
    print("   ✅ I列复制完成")
    
    # ============================================================
    # 【第四步】3.3 K25 = E5 + " " + (从K26提取并转换的 Cx.x)
    # ============================================================
    e5_value = ws.cell(row=5, column=5).value
    k26_raw = ws.cell(row=26, column=11).value
    
    cx_x = extract_cx_x_from_string(k26_raw)
    if not cx_x:
        f9_value = ws.cell(row=9, column=6).value
        cx_x = extract_cx_x_from_string(f9_value) if f9_value else ""
    
    k25_value = f"{e5_value} {cx_x}" if e5_value and cx_x else ""
    ws.cell(row=25, column=11, value=k25_value)
    print(f"   ✅ K25 = {k25_value}")
    
    ws.cell(row=158, column=11, value=k25_value)
    print(f"   ✅ K158 = {k25_value}")
    ws.cell(row=159, column=11, value=k26_raw)
    print(f"   ✅ K159 = {k26_raw}")
    
    # ============================================================
    # 【第五步】步骤10：填充 D9、F9、H9、K11
    # ============================================================
    print("📝 步骤10：填充 D9、F9、H9、K11...")
    ws.cell(row=9, column=4, value=阶段)
    ws.cell(row=9, column=6, value=Config)
    ws.cell(row=9, column=8, value=报告类型)
    ws.cell(row=11, column=11, value=机台号)
    
    print(f"   ✅ D9 = {阶段}")
    print(f"   ✅ F9 = {Config}")
    print(f"   ✅ H9 = {报告类型}")
    print(f"   ✅ K11 = {机台号}")

    # ============================================================
    # 【第六步】I26 = F9 + " " + H9
    # ============================================================
    f9_value = ws.cell(row=9, column=6).value
    h9_value = ws.cell(row=9, column=8).value
    i26_value = f"{f9_value} {h9_value}" if f9_value and h9_value else ""
    ws.cell(row=26, column=9, value=i26_value)
    print(f"   ✅ I26 = {i26_value}")
    
    # ============================================================
    # 【第七步】H4更新日期
    # ============================================================
    ws.cell(row=4, column=8, value=today_slash)
    print(f"   ✅ H4 = {today_slash}")
    
    # ============================================================
    # 【第八步】P5 - 从ERS文件获取版本号
    # ============================================================
    ers_file = get_first_file_in_folder(os.path.join(data_folder_path, "ERS"))
    if ers_file:
        ers_rev = extract_revision_from_file(ers_file, r'Table : Process Control Rev ([0-9]+\.[0-9]+)')
        if ers_rev:
            ws.cell(row=5, column=16, value=ers_rev)
            print(f"   ✅ P5 = {ers_rev}")
        else:
            print(f"   ⚠️ 未在ERS文件中找到版本号")
    else:
        print(f"   ⚠️ ERS文件夹为空或不存在")
    
    # ============================================================
    # 【第九步】Q5 - 从ERS文件名提取版本号
    # ============================================================
    if ers_file:
        ers_filename = os.path.basename(ers_file)
        match = re.search(r'Rev([0-9]+\.[0-9]+)', ers_filename)
        if match:
            ws.cell(row=5, column=17, value=match.group(1))
            print(f"   ✅ Q5 = {match.group(1)}")
        else:
            print(f"   ⚠️ 未从ERS文件名提取到版本号：{ers_filename}")
    
    # ============================================================
    # 【第十步】R5 - 从VSR文件名提取版本号
    # ============================================================
    vsr_file = get_first_file_in_folder(os.path.join(data_folder_path, "VSR"))
    if vsr_file:
        vsr_filename = os.path.basename(vsr_file)
        match = re.search(r'Rev\s*([0-9]+)', vsr_filename)
        if match:
            ws.cell(row=5, column=18, value=match.group(1))
            print(f"   ✅ R5 = {match.group(1)}")
        else:
            print(f"   ⚠️ 未从VSR文件名提取到版本号：{vsr_filename}")
    
    # ============================================================
    # 【第十一步】S5 - 从MCO文件名提取版本号
    # ============================================================
    mco_file = get_first_file_in_folder(os.path.join(data_folder_path, "MCO"))
    if mco_file:
        mco_filename = os.path.basename(mco_file)
        match = re.search(r'[0-9]+-[0-9]+-([0-9]+)', mco_filename)
        if match:
            ws.cell(row=5, column=19, value=match.group(1))
            print(f"   ✅ S5 = {match.group(1)}")
        else:
            parts = mco_filename.split('-')
            if len(parts) >= 3:
                val = parts[2].split('.')[0]
                ws.cell(row=5, column=19, value=val)
                print(f"   ✅ S5 = {val}")
            else:
                print(f"   ⚠️ 未从MCO文件名提取到版本号：{mco_filename}")
    
    # ============================================================
    # 【第十二步】更新 Raw Data Sheet
    # ============================================================
    update_raw_data_report(wb, data_folder_path, selected_fixtures)

    # ============================================================
    # 【第十三步】更新 DCR Sheet
    # ============================================================
    update_dcr_report(wb, data_folder_path, selected_fixtures)
    # ============================================================
    # 【第十四步】更新 Weld Spot Diameter Sheet
    # ============================================================
    update_weld_spot_diameter_report(wb, data_folder_path, selected_fixtures)
    # ============================================================
    # 【第十五步】更新 Bond strength Sheet
    # ============================================================
    update_bond_strength_report(wb, data_folder_path, selected_fixtures, 报告类型)
    # ============================================================
    # 【第十六步】更新 Outgassing Sheet
    # ============================================================
    update_outgassing_report(wb, data_folder_path, selected_fixtures)
    # ============================================================
    # 【第十七步】更新 Visual Inspection Sheet
    # ============================================================
    update_visual_inspection_report(wb, data_folder_path, selected_fixtures)
    # ============================================================
    # 更新 X-Section Sheet
    # ============================================================
    update_x_section_report(wb, data_folder_path, selected_fixtures)
    # ============================================================
    # 保存文件
    # ============================================================
    wb.save(output_path)
    print("=" * 60)
    print(f"✅ 报告更新完成！")
    print(f"   📄 文件位置：{output_path}")
    print("=" * 60)
    
    return output_path


# --- 使用示例 ---
if __name__ == "__main__":
    import tkinter as tk
    from tkinter import filedialog, messagebox
    import sys

    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)

    # 1. 选择数据文件夹（命名格式：专案 阶段 Config 机台号 M/PBO）
    data_folder = filedialog.askdirectory(
        title="请选择数据文件夹（命名格式：专案 阶段 Config 机台号 M/PBO）"
    )
    if not data_folder:
        print("❌ 未选择数据文件夹，程序退出")
        sys.exit(0)

    # 2. 选择报告模板文件（.xlsx）
    template_file = filedialog.askopenfilename(
        title="请选择报告模板文件（.xlsx格式）",
        filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
    )
    if not template_file:
        print("❌ 未选择模板文件，程序退出")
        sys.exit(0)

    # 3. 验证数据文件夹名称是否符合格式
    folder_name = os.path.basename(data_folder)
    if not parse_folder_name(folder_name):
        messagebox.showerror(
            "格式错误",
            f"数据文件夹名称格式不正确！\n应为：专案 阶段 Config 机台号 M/PBO\n当前：{folder_name}"
        )
        sys.exit(1)

    # 4. 输出文件夹默认与数据文件夹相同（可自行修改）
    output_folder = None  # 设为 None 表示与数据文件夹相同

    # 5. 执行更新
    if os.path.exists(data_folder) and os.path.exists(template_file):
        update_summary_report(template_file, data_folder, output_folder)
    else:
        print("❌ 错误：找不到数据文件夹或模板文件")
        sys.exit(1)